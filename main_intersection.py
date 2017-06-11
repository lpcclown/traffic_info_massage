import csv
import glob
import datetime
import openpyxl
from openpyxl import load_workbook
import xlsxwriter
import xlrd
from BaseHTTPServer import BaseHTTPRequestHandler, HTTPServer
import cgi
import os
import shutil
from win32com.client import Dispatch

def pdf_to_csv(filename, separator, threshold):
    from cStringIO import StringIO
    from pdfminer.converter import LTChar, TextConverter
    from pdfminer.layout import LAParams
    from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
    from pdfminer.pdfpage import PDFPage

    class CsvConverter(TextConverter):
        def __init__(self, *args, **kwargs):
            TextConverter.__init__(self, *args, **kwargs)
            self.separator = separator
            self.threshold = threshold

        def end_page(self, i):
            from collections import defaultdict
            lines = defaultdict(lambda: {})
            for child in self.cur_item._objs:  # <-- changed
                if isinstance(child, LTChar):
                    (_, _, x, y) = child.bbox
                    line = lines[int(-y)]
                    line[x] = child._text.encode(self.codec)  # <-- changed
            for y in sorted(lines.keys()):
                line = lines[y]
                self.line_creator(line)
                self.outfp.write(self.line_creator(line))
                self.outfp.write("\n")

        def line_creator(self, line):
            keys = sorted(line.keys())
            # calculate the average distange between each character on this row
            average_distance = sum([keys[i] - keys[i - 1] for i in range(1, len(keys))]) / len(keys)
            # append the first character to the result
            result = [line[keys[0]]]
            for i in range(1, len(keys)):
                # if the distance between this character and the last character is greater than the average*threshold
                if (keys[i] - keys[i - 1]) > average_distance * self.threshold:
                    # append the separator into that position
                    result.append(self.separator)
                # append the character
                result.append(line[keys[i]])
            printable_line = ''.join(result)
            return printable_line

    # ... the following part of the code is a remix of the
    # convert() function in the pdfminer/tools/pdf2text module
    rsrc = PDFResourceManager()
    outfp = StringIO()
    device = CsvConverter(rsrc, outfp, codec="utf-8", laparams=LAParams())
    # becuase my test documents are utf-8 (note: utf-8 is the default codec)

    fp = open(filename, 'rb')

    interpreter = PDFPageInterpreter(rsrc, device)
    for i, page in enumerate(PDFPage.get_pages(fp)):
        outfp.write("START PAGE %d\n" % i)
        if page is not None:
            print 'none'
            interpreter.process_page(page)
        outfp.write("END PAGE %d\n" % i)

    device.close()
    fp.close()

    return outfp.getvalue()


if __name__ == '__main__':
    # the separator to use with the CSV
    separator = '|'
    # the distance multiplier after which a character is considered part of a new word/column/block. Usually 1.5 works quite well
    threshold = 1.5
    # print pdf_to_csv('Sample MOE Plan Report_Avg Grn per Intersection_MU (4).pdf', separator, threshold)

    duration = 3600  # need to be grab from website page
    MOE_plan_pdf = 'Sample MOE Plan Report_Avg Grn per Intersection_MU (4).pdf'
    TOD_schedule_pdf = '3129 - TODSchedule_MU.pdf'

    # Starting for V1 second part
    def calculate_ending_time(start_time):
        adding_hour = int(duration / 60)
        adding_minute = int(duration % 60)
        ending_time_hour = int(str(start_time).split(':')[0]) + adding_hour
        if ending_time_hour < 10:
            ending_time_hour = "0" + str(ending_time_hour)
        ending_time_minute = int(str(start_time).split(':')[1]) + adding_minute
        if ending_time_minute < 10:
            ending_time_minute = "0" + str(ending_time_minute)
        return str(ending_time_hour) + ":" + str(ending_time_minute)


    def transform_time(original_string):
        if len(original_string.split(' ')) > 2:
            old_hour_minute = original_string.split(' ', 5)[4].rsplit(':', 1)[0]
            old_hour = old_hour_minute.split(':', 1)[0]
            old_date = original_string.split(' ', 5)[3] + original_string.split(' ', 5)[1] + \
                       original_string.split(' ', 5)[
                           2]
            date = original_string[0:24]

            new_minute = int(original_string.split(' ', 5)[4].split(':', 2)[1]) + 5
            if new_minute < 10:
                new_minute = "0" + str(new_minute)
            return str(old_hour_minute) + " - " + str(old_hour) + ":" + str(new_minute)


    def transfrom_date(original_string):
        old_hour_minute = original_string.split(' ', 5)[4].rsplit(':', 1)[0]
        old_hour = old_hour_minute.split(':', 1)[0]
        old_date = original_string.split(' ', 5)[3] + original_string.split(' ', 5)[1] + original_string.split(' ', 5)[
            2]
        date = original_string[0:24]
        return str(datetime.datetime.strptime(date, "%a %b %d %Y %H:%M:%S")).split(' ')[0]


    def compare_time(input_time, interval_upper_time):
        input_time_hour = str(input_time).split(':', 0)
        input_time_minute = str(input_time).split(':', 1)
        interval_upper_time_hour = str(interval_upper_time).split(':', 0)
        interval_upper_time_minute = str(interval_upper_time).split(':', 1)
        if input_time_hour > interval_upper_time_hour:
            return 0  # no need to write
            if input_time_minute > interval_upper_time_minute:
                return 0  # no need to write
        else:
            return 1


    # here will be three input files instead of one in V1
    input_path = 'D:\Users\LIU\PycharmProjects\\bluetooth_V2\\'

    travel_time_xml_file = 'Intersection Approach Travel Time_ 67 Ave at 103 St (2).xml'
    # need to rename the xml file, other wise the Workbooks.open is not working.
    # TODO, the further reason about why original file name is not working
    renamed_travel_time_xml_file = 'Travel_Time.xml'
    shutil.copyfile(travel_time_xml_file, renamed_travel_time_xml_file)
    travel_time_excel_file = 'Travel_Time'  # suffix will be added by function SaveAs
    xlApp = Dispatch("Excel.Application")
    xlBook = xlApp.Workbooks.open(input_path + renamed_travel_time_xml_file)
    xlBook.SaveAs(input_path + travel_time_excel_file, FileFormat=51)
    xlBook.Close(SaveChanges=0)
    # Read content from excel
    wb = load_workbook(filename=travel_time_excel_file + '.xlsx')
    sheet_ranges = wb['Travel time']
    print(sheet_ranges['C6'].value)

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "DataSheet"
    date_sheet_name = sheet.title
    sheet['A1'] = 'hello world'

    result = {}
    a = ""
    b = ""
    c = 1
    d = 1
    common_array = []


    def init_final_report():
        sheet["A1"] = "Time"
        sheet["B1"] = "Period"
        sheet["C1"] = "NT"
        sheet["D1"] = "Time Interval (NT)"
        sheet["E1"] = "Travel Time (NT)"
        sheet["F1"] = "ST"
        sheet["G1"] = "Time Interval (ST)"
        sheet["H1"] = "Travel Time (ST)"
        sheet["I1"] = "ET"
        sheet["J1"] = "Time Interval (ET)"
        sheet["K1"] = "Travel Time (ET)"
        sheet["L1"] = "WT"
        sheet["M1"] = "Time Interval (WT)"
        sheet["N1"] = "Travel Time (WT)"

    # Fill out column headers
    init_final_report()

    # Fill out column Time
    var = 1
    date_column = 'C'
    time_row = 2
    date_index = 0
    while var == 1:  # This constructs an infinite loop to check how many days data
        date_item = sheet_ranges[date_column + '5'].value
        if date_item != '':
            # print date_item
            date_column = chr(ord(date_column) + 1)
            sheet['A' + str(time_row)] = date_item
            date_index += 1
            # TODO 1440 should be changed according to user selected time duration
            time_row = 2 + 1440 * date_index
        else:
            break

    # Fill out column Period
    # TODO to discuss with Boyuan what is the content to be filled in

    # Fill out columns: NT, ST, ET, WT columns name: 'C' 'F' 'I' 'L'
    # TODO this column is involving other two pdf files
    MOE_plan = pdf_to_csv(MOE_plan_pdf, separator, threshold)
    TOD_schedule = pdf_to_csv(TOD_schedule_pdf, separator, threshold)
    TOD_schedule_lines = TOD_schedule.split('\n')
    directions = TOD_schedule_lines[11].split('|')
    MOE_plan_lines = MOE_plan.split('\n')
    direction_index = 0
    directions_and_avggreens = []
    for MOE_plan_lines_data_zone in MOE_plan_lines[8: 16]:
        direction = directions[direction_index]
        if direction[-1:] == 'T':
            directions_and_avggreens.append(direction + ','+ MOE_plan_lines_data_zone.split('|')[2])
        direction_index += 1
    total_lines = 4320  # TODO this total line number should be calculated based on users input
    for direction_and_avggreen in directions_and_avggreens:
        if direction_and_avggreen.split(',')[0] == 'NBT':
            for line in range(2, total_lines + 2):
                sheet['C' + str(line)] = direction_and_avggreen.split(',')[1]
        if direction_and_avggreen.split(',')[0] == 'SBT':
            for line in range(2, total_lines + 2):
                sheet['F' + str(line)] = direction_and_avggreen.split(',')[1]
        if direction_and_avggreen.split(',')[0] == 'EBT':
            for line in range(2, total_lines + 2):
                sheet['I' + str(line)] = direction_and_avggreen.split(',')[1]
        if direction_and_avggreen.split(',')[0] == 'WBT':
            for line in range(2, total_lines + 2):
                sheet['L' + str(line)] = direction_and_avggreen.split(',')[1]

    # Fill out column Time Interval (NT), Time Interval (ST), Time Interval (ET), Time Interval (WT),
    # columns name 'D' 'G' 'J' 'M'
    source_time_interval_column = 'A'
    minute_row = 2
    for date_travel_time in range(0, date_index):  # Loop different date's travel time data
        input_minute_row = 6
        while var == 1:  # This constructs an infinite loop to check how many minute data in one day
            time_interval = sheet_ranges[source_time_interval_column + str(input_minute_row)].value
            input_minute_row += 1
            if time_interval is not None:
                sheet['D' + str(minute_row)] = time_interval
                sheet['G' + str(minute_row)] = time_interval
                sheet['J' + str(minute_row)] = time_interval
                sheet['M' + str(minute_row)] = time_interval
                minute_row += 1
            else:
                break

    # Fill out column Travel Time (NT), column 'E'
    nt_travel_time_column = 'C'
    time_travel_row = 2
    for date_travel_time in range(0, date_index):  # Loop different date's travel time data
        input_travel_time_row = 6
        while var == 1:  # This constructs an infinite loop to check how many travel time data in one day
            travel_time = sheet_ranges[nt_travel_time_column + str(input_travel_time_row)].value
            input_travel_time_row += 1
            if travel_time is not None:
                sheet['E' + str(time_travel_row)] = travel_time
                time_travel_row += 1
            else:
                break
        nt_travel_time_column = chr(ord(nt_travel_time_column) + 1)
    # Fill out column Travel Time (ST), column 'H'
    st_travel_time_column = chr(ord('C') + 1 * (date_index+1))
    time_travel_row = 2
    for date_travel_time in range(0, date_index):  # Loop different date's travel time data
        input_travel_time_row = 6
        while var == 1:  # This constructs an infinite loop to check how many travel time data in one day
            travel_time = sheet_ranges[st_travel_time_column + str(input_travel_time_row)].value
            input_travel_time_row += 1
            if travel_time is not None:
                sheet['H' + str(time_travel_row)] = travel_time
                time_travel_row += 1
            else:
                break
        st_travel_time_column = chr(ord(st_travel_time_column) + 1)
    # Fill out column Travel Time (ET), column 'K'
    et_travel_time_column = chr(ord('C') + 2 * (date_index+1))
    time_travel_row = 2
    for date_travel_time in range(0, date_index):  # Loop different date's travel time data
        input_travel_time_row = 6
        while var == 1:  # This constructs an infinite loop to check how many travel time data in one day
            travel_time = sheet_ranges[et_travel_time_column + str(input_travel_time_row)].value
            input_travel_time_row += 1
            if travel_time is not None:
                sheet['K' + str(time_travel_row)] = travel_time
                time_travel_row += 1
            else:
                break
        et_travel_time_column = chr(ord(et_travel_time_column) + 1)
    # Fill out column Travel Time (WT), column 'N'
    wt_travel_time_column = chr(ord('C') + 3 * (date_index+1))
    time_travel_row = 2
    for date_travel_time in range(0, date_index):  # Loop different date's travel time data
        input_travel_time_row = 6
        while var == 1:  # This constructs an infinite loop to check how many travel time data in one day
            travel_time = sheet_ranges[wt_travel_time_column + str(input_travel_time_row)].value
            input_travel_time_row += 1
            if travel_time is not None:
                sheet['N' + str(time_travel_row)] = travel_time
                time_travel_row += 1
            else:
                break
        wt_travel_time_column = chr(ord(wt_travel_time_column) + 1)

    book.save('bluetooth_V2_intersection_output.xlsx')

    wb.close()

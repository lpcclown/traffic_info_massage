import csv
import glob
import datetime
import openpyxl
import xlsxwriter
import xlrd
from BaseHTTPServer import BaseHTTPRequestHandler, HTTPServer
import cgi


class Handler(BaseHTTPRequestHandler):
    def do_GET(self):
        self.send_response(200)
        self.end_headers()
        self.wfile.write("""
            <html><head></head>
            <body>
            <form method="POST">
            Starting Time:
            <input type="text" name="startingtime" value="03:10">
            </br>
            Duration:
            <input type="text" name="duration" value="150">
            </br>
            SSID:
            <input type="text" name="ssid" value="4869">
            </br>
            <input type="submit" name="submit" value="submit">
            </form>
            </body>
            </html>
            """)
        return

    def do_POST(self):
        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={'REQUEST_METHOD': 'POST',
                     'CONTENT_TYPE': self.headers['Content-Type'],
                     })
        starting_time_dic = form["startingtime"]
        starting_time = starting_time_dic.value
        duration_dic = form["duration"]
        duration = int(duration_dic.value)
        ssid = form["ssid"].value.split(",")
        movement_head = ['WT', 'ET', 'NT', 'ST', 'WL', 'EL', 'NL', 'SL']
        another_final_file = "finalReport_" + ssid[0] + "_6_30_360_.csv"
        excel_file_name = "bluetooth_output.xlsx"

        def calculate_ending_time(start_time):
            adding_hour = int(duration / 60)
            adding_minute = int(duration % 60)
            ending_time_hour = int(str(start_time).split(':')[0]) + adding_hour
            if ending_time_hour < 10:
                ending_time_hour = "0" + str(ending_time_hour)
            ending_time_minute = int(str(start_time).split(':')[1]) + adding_minute
            if ending_time_minute < 10:
                ending_time_minute = "0" + str(ending_time_minute)
            return str(ending_time_hour) + ":" + str(ending_time_minute)


        def transform_time(original_string):
            if len(original_string.split(' ')) > 2:
                old_hour_minute = original_string.split(' ', 5)[4].rsplit(':', 1)[0]
                old_hour = old_hour_minute.split(':', 1)[0]
                old_date = original_string.split(' ', 5)[3] + original_string.split(' ', 5)[1] + original_string.split(' ', 5)[
                    2]
                date = original_string[0:24]

                new_minute = int(original_string.split(' ', 5)[4].split(':', 2)[1]) + 5
                if new_minute < 10:
                    new_minute = "0" + str(new_minute)
                return str(old_hour_minute) + " - " + str(old_hour) + ":" + str(new_minute)


        def transfrom_date(original_string):
            old_hour_minute = original_string.split(' ', 5)[4].rsplit(':', 1)[0]
            old_hour = old_hour_minute.split(':', 1)[0]
            old_date = original_string.split(' ', 5)[3] + original_string.split(' ', 5)[1] + original_string.split(' ', 5)[2]
            date = original_string[0:24]
            return str(datetime.datetime.strptime(date, "%a %b %d %Y %H:%M:%S")).split(' ')[0]


        def compare_time(input_time, interval_upper_time):
            input_time_hour = str(input_time).split(':', 0)
            input_time_minute = str(input_time).split(':', 1)
            interval_upper_time_hour = str(interval_upper_time).split(':', 0)
            interval_upper_time_minute = str(interval_upper_time).split(':', 1)
            if input_time_hour > interval_upper_time_hour:
                return 0  # no need to write
                if input_time_minute > interval_upper_time_minute:
                    return 0  # no need to write
            else:
                return 1

        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = "DataSheet"
        date_sheet_name = sheet.title
        sheet['A1'] = 'hello world'

        result = {}
        a = ""
        b = ""
        c = 1
        d = 1
        common_array = []

        def init_final_report():
            sheet["A1"] = "Time"
            sheet["B1"] = "Period"
            sheet["C1"] = "WT"
            sheet["D1"] = "Time Interval (WT)"
            sheet["E1"] = "Travel Time (WT)"
            sheet["F1"] = "ET"
            sheet["G1"] = "Time Interval (ET)"
            sheet["H1"] = "Travel Time (ET)"
            sheet["I1"] = "NT"
            sheet["J1"] = "Time Interval (NT)"
            sheet["K1"] = "Travel Time (NT)"
            sheet["L1"] = "ST"
            sheet["M1"] = "Time Interval (ST)"
            sheet["N1"] = "Travel Time (ST)"
            sheet["O1"] = "WL"
            sheet["P1"] = "Time Interval (WL)"
            sheet["Q1"] = "Travel Time (WL)"
            sheet["R1"] = "EL"
            sheet["S1"] = "Time Interval (EL)"
            sheet["T1"] = "Travel Time (EL)"
            sheet["U1"] = "NL"
            sheet["V1"] = "Time Interval (NL)"
            sheet["W1"] = "Travel Time (NL)"
            sheet["X1"] = "SL"
            sheet["Y1"] = "Time Interval (SL)"
            sheet["Z1"] = "Travel Time (SL)"


        def fill_empty(column_name, total_row_number):
            i = 2
            while (i < total_row_number):
                if sheet[column_name + str(i)].value is None:
                    sheet[column_name + str(i)].value = sheet[column_name + str(i - 1)].value
                i += 1


        for strSSID in ssid:
            for infile in sorted(glob.glob('*.csv')):
                filename = str(infile)
                if strSSID in filename and "WT" in filename:
                    c = 1
                    print "Current File Being Processed is: " + filename
                    reader = csv.DictReader(open(filename))
                    for row in reader:
                        if len(str(row).split(',')) > 3:  # ignore missing value dirty rows
                            try:
                                for column, value in sorted(row.iteritems(), reverse=True):
                                    if column == "Time\"":
                                        if (compare_time(starting_time,
                                                         transform_time(str(value)).split(" - ")[1]) == 1 and compare_time(
                                            transform_time(str(value)).split(" - ")[0],
                                                calculate_ending_time(starting_time)) == 1):
                                            c += 1
                                            cell_name = "D" + str(c)
                                            sheet[cell_name] = transform_time(str(value))
                                            cell_previous_name = "A" + str(int(c) - 1)
                                            cell_name = "A" + str(c)
                                            if len(str(sheet[cell_previous_name].value)) > 5:  # used to compare None
                                                temp = str(sheet[cell_previous_name].value)
                                            if transfrom_date(str(value)) != temp:
                                                sheet[cell_name] = transfrom_date(str(value))
                                        else:
                                            break
                                    if column == "Strength":
                                        cell_name = "E" + str(c)
                                        sheet[cell_name] = datetime.datetime.strptime(value, '%M:%S')
                                        sheet[cell_name].number_format = "mm:ss"

                            except:
                                print("There is dirty data row in the file.")
                if strSSID in filename and "ET" in filename:
                    c = 1
                    print "Current File Being Processed is: " + filename
                    reader = csv.DictReader(open(filename))
                    for row in reader:
                        if len(str(row).split(',')) > 3:  # ignore missing value dirty rows
                            try:
                                for column, value in sorted(row.iteritems(), reverse=True):
                                    if column == "Time\"":
                                        if (compare_time(starting_time,
                                                         transform_time(str(value)).split(" - ")[1]) == 1 and compare_time(
                                            transform_time(str(value)).split(" - ")[0],
                                                calculate_ending_time(starting_time)) == 1):
                                            c += 1
                                            cell_name = "G" + str(c)
                                            sheet[cell_name] = transform_time(str(value))
                                            cell_previous_name = "A" + str(int(c) - 1)
                                            cell_name = "A" + str(c)
                                            if len(str(sheet[cell_previous_name].value)) > 5:  # used to compare None
                                                temp = str(sheet[cell_previous_name].value)
                                            if transfrom_date(str(value)) != temp:
                                                sheet[cell_name] = transfrom_date(str(value))
                                                common_array.append(c)
                                                line = c  # used to generate chart
                                        else:
                                            break
                                    if column == "Strength":
                                        cell_name = "H" + str(c)
                                        sheet[cell_name] = datetime.datetime.strptime(value, '%M:%S')
                                        sheet[cell_name].number_format = "mm:ss"
                            except:
                                print("There is dirty data row in the file.")
                if strSSID in filename and "NT" in filename:
                    c = 1
                    print "Current File Being Processed is: " + filename
                    reader = csv.DictReader(open(filename))
                    for row in reader:
                        if len(str(row).split(',')) > 3:  # ignore missing value dirty rows
                            try:
                                for column, value in sorted(row.iteritems(), reverse=True):
                                    if column == "Time\"":
                                        if (compare_time(starting_time,
                                                         transform_time(str(value)).split(" - ")[1]) == 1 and compare_time(
                                            transform_time(str(value)).split(" - ")[0],
                                                calculate_ending_time(starting_time)) == 1):
                                            c += 1
                                            cell_name = "J" + str(c)
                                            sheet[cell_name] = transform_time(str(value))
                                            cell_previous_name = "A" + str(int(c) - 1)
                                            cell_name = "A" + str(c)
                                            if len(str(sheet[cell_previous_name].value)) > 5:  # used to compare None
                                                temp = str(sheet[cell_previous_name].value)
                                            if transfrom_date(str(value)) != temp:
                                                sheet[cell_name] = transfrom_date(str(value))
                                        else:
                                            break
                                    if column == "Strength":
                                        cell_name = "K" + str(c)
                                        sheet[cell_name] = datetime.datetime.strptime(value, '%M:%S')
                                        sheet[cell_name].number_format = "mm:ss"
                            except:
                                print("There is dirty data row in the file.")
                if strSSID in filename and "ST" in filename:
                    c = 1
                    print "Current File Being Processed is: " + filename
                    reader = csv.DictReader(open(filename))
                    for row in reader:
                        if len(str(row).split(',')) > 3:  # ignore missing value dirty rows
                            try:
                                for column, value in sorted(row.iteritems(), reverse=True):
                                    if column == "Time\"":
                                        if (compare_time(starting_time,
                                                         transform_time(str(value)).split(" - ")[1]) == 1 and compare_time(
                                            transform_time(str(value)).split(" - ")[0],
                                                calculate_ending_time(starting_time)) == 1):
                                            c += 1
                                            cell_name = "M" + str(c)
                                            sheet[cell_name] = transform_time(str(value))
                                            cell_previous_name = "A" + str(int(c) - 1)
                                            cell_name = "A" + str(c)
                                            if len(str(sheet[cell_previous_name].value)) > 5:  # used to compare None
                                                temp = str(sheet[cell_previous_name].value)
                                            if transfrom_date(str(value)) != temp:
                                                sheet[cell_name] = transfrom_date(str(value))
                                        else:
                                            break
                                    if column == "Strength":
                                        cell_name = "N" + str(c)
                                        sheet[cell_name] = datetime.datetime.strptime(value, '%M:%S')
                                        sheet[cell_name].number_format = "mm:ss"
                            except:
                                print("There is dirty data row in the file.")
                if strSSID in filename and "WL" in filename:
                    c = 1
                    print "Current File Being Processed is: " + filename
                    reader = csv.DictReader(open(filename))
                    for row in reader:
                        if len(str(row).split(',')) > 3:  # ignore missing value dirty rows
                            try:
                                for column, value in sorted(row.iteritems(), reverse=True):
                                    if column == "Time\"":
                                        if (compare_time(starting_time,
                                                         transform_time(str(value)).split(" - ")[1]) == 1 and compare_time(
                                            transform_time(str(value)).split(" - ")[0],
                                                calculate_ending_time(starting_time)) == 1):
                                            c += 1
                                            cell_name = "P" + str(c)
                                            sheet[cell_name] = transform_time(str(value))
                                            cell_previous_name = "A" + str(int(c) - 1)
                                            cell_name = "A" + str(c)
                                            if len(str(sheet[cell_previous_name].value)) > 5:  # used to compare None
                                                temp = str(sheet[cell_previous_name].value)
                                            if transfrom_date(str(value)) != temp:
                                                sheet[cell_name] = transfrom_date(str(value))
                                        else:
                                            break
                                    if column == "Strength":
                                        cell_name = "Q" + str(c)
                                        sheet[cell_name] = datetime.datetime.strptime(value, '%M:%S')
                                        sheet[cell_name].number_format = "mm:ss"
                            except:
                                print("There is dirty data row in the file.")
                if strSSID in filename and "EL" in filename:
                    c = 1
                    print "Current File Being Processed is: " + filename
                    reader = csv.DictReader(open(filename))
                    for row in reader:
                        if len(str(row).split(',')) > 3:  # ignore missing value dirty rows
                            try:
                                for column, value in sorted(row.iteritems(), reverse=True):
                                    if column == "Time\"":
                                        if (compare_time(starting_time,
                                                         transform_time(str(value)).split(" - ")[1]) == 1 and compare_time(
                                            transform_time(str(value)).split(" - ")[0],
                                                calculate_ending_time(starting_time)) == 1):
                                            c += 1
                                            cell_name = "S" + str(c)
                                            sheet[cell_name] = transform_time(str(value))
                                            cell_previous_name = "A" + str(int(c) - 1)
                                            cell_name = "A" + str(c)
                                            if len(str(sheet[cell_previous_name].value)) > 5:  # used to compare None
                                                temp = str(sheet[cell_previous_name].value)
                                            if transfrom_date(str(value)) != temp:
                                                sheet[cell_name] = transfrom_date(str(value))
                                        else:
                                            break
                                    if column == "Strength":
                                        cell_name = "T" + str(c)
                                        sheet[cell_name] = datetime.datetime.strptime(value, '%M:%S')
                                        sheet[cell_name].number_format = "mm:ss"
                            except:
                                print("There is dirty data row in the file.")

                if strSSID in filename and "NL" in filename:
                    c = 1
                    print "Current File Being Processed is: " + filename
                    reader = csv.DictReader(open(filename))
                    for row in reader:
                        if len(str(row).split(',')) > 3:  # ignore missing value dirty rows
                            try:
                                for column, value in sorted(row.iteritems(), reverse=True):
                                    if column == "Time\"":
                                        if (compare_time(starting_time,
                                                         transform_time(str(value)).split(" - ")[1]) == 1 and compare_time(
                                            transform_time(str(value)).split(" - ")[0],
                                                calculate_ending_time(starting_time)) == 1):
                                            c += 1
                                            cell_name = "V" + str(c)
                                            sheet[cell_name] = transform_time(str(value))
                                            cell_previous_name = "A" + str(int(c) - 1)
                                            cell_name = "A" + str(c)
                                            if len(str(sheet[cell_previous_name].value)) > 5:  # used to compare None
                                                temp = str(sheet[cell_previous_name].value)
                                            if transfrom_date(str(value)) != temp:
                                                sheet[cell_name] = transfrom_date(str(value))
                                        else:
                                            break
                                    if column == "Strength":
                                        cell_name = "W" + str(c)
                                        sheet[cell_name] = datetime.datetime.strptime(value, '%M:%S')
                                        sheet[cell_name].number_format = "mm:ss"
                            except:
                                print("There is dirty data row in the file.")
                if strSSID in filename and "SL" in filename:
                    c = 1
                    print "Current File Being Processed is: " + filename
                    reader = csv.DictReader(open(filename))
                    for row in reader:
                        if len(str(row).split(',')) > 3:  # ignore missing value dirty rows
                            try:
                                for column, value in sorted(row.iteritems(), reverse=True):
                                    if column == "Time\"":
                                        if (compare_time(starting_time,
                                                         transform_time(str(value)).split(" - ")[1]) == 1 and compare_time(
                                            transform_time(str(value)).split(" - ")[0],
                                                calculate_ending_time(starting_time)) == 1):
                                            c += 1
                                            cell_name = "Y" + str(c)
                                            sheet[cell_name] = transform_time(str(value))
                                            cell_previous_name = "A" + str(int(c) - 1)
                                            cell_name = "A" + str(c)
                                            if len(str(sheet[cell_previous_name].value)) > 5:  # used to compare None
                                                temp = str(sheet[cell_previous_name].value)
                                            if transfrom_date(str(value)) != temp:
                                                sheet[cell_name] = transfrom_date(str(value))
                                        else:
                                            break
                                    if column == "Strength":
                                        cell_name = "Z" + str(c)
                                        sheet[cell_name] = datetime.datetime.strptime(value, '%M:%S')
                                        sheet[cell_name].number_format = "mm:ss"
                            except:
                                print("There is dirty data row in the file.")

        reader = csv.DictReader(open(another_final_file))
        i = 0
        for row in reader:
            for column, value in sorted(row.iteritems()):
                # if i < c:
                if str(column) == "Period":
                    for common_row in common_array:
                        cell_name = "B" + str(common_row)
                        sheet[cell_name] = str(value)
                if str(column) == "WT":
                    cell_name = "C" + str(common_array[i])
                    sheet[cell_name] = float(value)
                if str(column) == "ET":
                    cell_name = "F" + str(common_array[i])
                    sheet[cell_name] = float(value)
                if str(column) == "NT":
                    cell_name = "I" + str(common_array[i])
                    sheet[cell_name] = float(value)
                if str(column) == "ST":
                    cell_name = "L" + str(common_array[i])
                    sheet[cell_name] = float(value)
                if str(column) == "WL":
                    cell_name = "O" + str(common_array[i])
                    sheet[cell_name] = float(value)
                if str(column) == "EL":
                    cell_name = "R" + str(common_array[i])
                    sheet[cell_name] = float(value)
                if str(column) == "NL":
                    cell_name = "U" + str(common_array[i])
                    sheet[cell_name] = float(value)
                if str(column) == "SL":
                    cell_name = "X" + str(common_array[i])
                    sheet[cell_name] = float(value)
            i += 1

        init_final_report()
        fill_empty("B", c + 1)
        fill_empty("C", c + 1)
        fill_empty("F", c + 1)
        fill_empty("I", c + 1)
        fill_empty("L", c + 1)
        fill_empty("O", c + 1)
        fill_empty("R", c + 1)
        fill_empty("U", c + 1)
        fill_empty("X", c + 1)

        book.save(excel_file_name)


        workbook = xlrd.open_workbook(excel_file_name)
        sheets = workbook.sheets()
        wb = xlsxwriter.Workbook(excel_file_name)
        time_format = wb.add_format({'num_format': 'm:ss'})

        for sheet in sheets:  # write data from old file
            newSheet = wb.add_worksheet(sheet.name)
            for row in range(sheet.nrows):
                for col in range(sheet.ncols):
                    if (col == 4 or col == 7 or col == 10 or col == 13
                            or col == 16 or col == 19 or col == 22 or col == 25):
                        newSheet.write(row, col, sheet.cell(row, col).value, time_format)
                    else:
                        newSheet.write(row, col, sheet.cell(row, col).value)


        def create_new_chart(workbook):
            chart = workbook.add_chart({'type': 'line'})

            chart.set_legend({'position': 'right'})

            # Add a chart title and some axis labels.
            chart.set_title({'name': 'Traffic Info'})
            chart.set_x_axis({'name': 'Time Intervals',})
            # chart.set_x_axis({'text_axis': True})
            chart.set_y_axis({'name': 'Bond', 'major_gridlines': {'visible': 0}})
            chart.set_y2_axis({'name': 'Travel Time'})
            return chart

        new_chart_sheet = wb.add_worksheet("ChartSheet")
        # new_chart_sheet = xfile.get_sheet_by_name('ChartSheet')
        i = 0
        j = 2


        for common_row in common_array:  # every loop draw 8 pic
            chart = create_new_chart(wb)
            chart.add_series({
                'name': date_sheet_name + '!$C$1',
                'values': (date_sheet_name + '!$C$2:$C$' + str(line)),
                'y2_axis': 1,
            })
            chart.add_series({
                'name': date_sheet_name + '!$E$1',
                'values': (date_sheet_name + '!$E$2:$E$' + str(line)),
            })
            chart.set_y_axis({'num_format': 'm:ss'})
            # Insert 1 the chart into the worksheet (with an offset).

            new_chart_sheet.insert_chart('B' + str(j), chart, {'x_offset': 25, 'y_offset': 10})

            chart = create_new_chart(wb)
            chart.add_series({
                'name': date_sheet_name + '!$F$1',
                'values': (date_sheet_name + '!$F$2:$F$' + str(line)),
                'y2_axis': 1,
            })
            chart.add_series({
                'name': date_sheet_name + '!$H$1',
                'values': (date_sheet_name + '!$H$2:$H$' + str(line)),
            })
            chart.set_y_axis({'num_format': 'm:ss'})
            # Insert 2 the chart into the worksheet (with an offset).
            new_chart_sheet.insert_chart('J' + str(j), chart, {'x_offset': 25, 'y_offset': 10})


            chart = create_new_chart(wb)
            chart.add_series({
                'name': date_sheet_name + '!$I$1',
                'values': (date_sheet_name + '!$I$2:$I$' + str(line)),
                'y2_axis': 1,
            })
            chart.add_series({
                'name': date_sheet_name + '!$K$1',
                'values': (date_sheet_name + '!$K$2:$K$' + str(line)),
            })
            chart.set_y_axis({'num_format': 'm:ss'})
            # Insert 3 the chart into the worksheet (with an offset).
            new_chart_sheet.insert_chart('R' + str(j), chart, {'x_offset': 25, 'y_offset': 10})

            chart = create_new_chart(wb)
            chart.add_series({
                'name': date_sheet_name + '!$L$1',
                'values': (date_sheet_name + '!$L$2:$L$' + str(line)),
                'y2_axis': 1,
            })
            chart.add_series({
                'name': date_sheet_name + '!$N$1',
                'values': (date_sheet_name + '!$N$2:$N$' + str(line)),
            })
            chart.set_y_axis({'num_format': 'm:ss'})
            # Insert 4 the chart into the worksheet (with an offset).
            new_chart_sheet.insert_chart('Z' + str(j), chart, {'x_offset': 25, 'y_offset': 10})


            chart = create_new_chart(wb)
            chart.add_series({
                'name': date_sheet_name + '!$O$1',
                'values': (date_sheet_name + '!$O$2:$O$' + str(line)),
                'y2_axis': 1,
            })
            chart.add_series({
                'name': date_sheet_name + '!$Q$1',
                'values': (date_sheet_name + '!$Q$2:$Q$' + str(line)),
            })
            chart.set_y_axis({'num_format': 'm:ss'})
            # Insert 5 the chart into the worksheet (with an offset).
            new_chart_sheet.insert_chart('AH' + str(j), chart, {'x_offset': 25, 'y_offset': 10})


            chart = create_new_chart(wb)
            chart.add_series({
                'name': date_sheet_name + '!$R$1',
                'values': (date_sheet_name + '!$R$2:$R$' + str(line)),
                'y2_axis': 1,
            })
            chart.add_series({
                'name': date_sheet_name + '!$T$1',
                'values': (date_sheet_name + '!$T$2:$T$' + str(line)),
            })
            chart.set_y_axis({'num_format': 'm:ss'})
            # Insert 6 the chart into the worksheet (with an offset).
            new_chart_sheet.insert_chart('AP' + str(j), chart, {'x_offset': 25, 'y_offset': 10})


            chart = create_new_chart(wb)
            chart.add_series({
                'name': date_sheet_name + '!$U$1',
                'values': (date_sheet_name + '!$U$2:$U$' + str(line)),
                'y2_axis': 1,
            })
            chart.add_series({
                'name': date_sheet_name + '!$W$1',
                'values': (date_sheet_name + '!$W$2:$W$' + str(line)),
            })
            chart.set_y_axis({'num_format': 'm:ss'})
            # Insert 7 the chart into the worksheet (with an offset).
            new_chart_sheet.insert_chart('AX' + str(j), chart, {'x_offset': 25, 'y_offset': 10})


            chart = create_new_chart(wb)
            chart.add_series({
                'name': date_sheet_name + '!$X$1',
                'values': (date_sheet_name + '!$X$2:$X$' + str(line)),
                'y2_axis': 1,
            })
            chart.add_series({
                'name': date_sheet_name + '!$Z$1',
                'values': (date_sheet_name + '!$Z$2:$Z$' + str(line)),
            })
            chart.set_y_axis({'num_format': 'm:ss'})
            # Insert 8 the chart into the worksheet (with an offset).
            new_chart_sheet.insert_chart('BF' + str(j), chart, {'x_offset': 25, 'y_offset': 10})

            i += 1
            j += 17  # for next row of pics


        wb.close()

        self.wfile.write("The output file is named bluetooth_output.xlsx")

        return
server = HTTPServer(('', 8181), Handler)
server.serve_forever()

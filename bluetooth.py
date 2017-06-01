import csv
import glob
import datetime
import openpyxl
import xlsxwriter
import xlrd
from BaseHTTPServer import BaseHTTPRequestHandler, HTTPServer
import cgi
import os
import shutil


class Handler(BaseHTTPRequestHandler):
    def do_GET(self):
        self.send_response(200)
        self.end_headers()
        self.wfile.write("""
            <html lang="en">
              <head>
                <title>Data Processing Portal</title>
                <!-- Required meta tags -->
                <meta charset="utf-8">
                <meta name="viewport" content="width=device-width, initial-scale=1, shrink-to-fit=no">
            
                <!-- Bootstrap CSS -->
                <link rel="stylesheet" href="https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0-alpha.6/css/bootstrap.min.css" integrity="sha384-rwoIResjU2yc3z8GV/NPeZWAv56rSmLldC3R/AZzGRnGxQQKnKkoFVhFQhNUwEyJ" crossorigin="anonymous">
                <style>
                    body{
                        margin:50px;
                    }
                    h1{
                        text-align:center;
                    }
                </style>
              </head>
              <body>
                <h1>Data Processing Portal</h1>
            
                <form method="POST">
                  <div class="form-group">
                    <label for="txtSSID">SSID(s):</label>
                    <input class="form-control" id="txtSSID" name="ssid" aria-describedby="ssidHelp" placeholder="Please enter SSID">
                    <small id="ssidHelp" class="form-text text-muted">Please enter SSID in the format like: <i>4869,4563</i></small>
                  </div>
                  <div class="form-group">
                    <label for="txtDates">Date(s):</label>
                    <input class="form-control" id="txtDates" name="date" aria-describedby="datesHelp"  placeholder="Please enter Date">
                    <small id="datesHelp" class="form-text text-muted">Please enter Dates in the format like: <i>20170213,20170214</i></small>
                  </div>
                  <div class="form-group">
                    <label for="txtStartTime">Starting Time:</label>
                    <input class="form-control" id="txtStartTime" name="startingtime" aria-describedby="startTimeHelp"  placeholder="Please enter Starting Time">
                    <small id="startTimeHelp" class="form-text text-muted">Please enter Starting Time with 24-hour clock by using this format: <i>06:30</i></small>
                  </div>
                  <div class="form-group">
                    <label for="txtTimeSpan">Time Span:</label>
                    <input class="form-control" id="txtTimeSpan" name="duration" aria-describedby="timeSpanHelp"  placeholder="Please enter Time Span">
                    <small id="timeSpanHelp" class="form-text text-muted">Please enter Time Span in minutes by using this format: <i>100</i></small>
                  </div>
                 
                  <button type="submit" name="submit" class="btn btn-primary btn-block">Submit</button>
                </form>
                <!-- jQuery first, then Tether, then Bootstrap JS. -->
                <script src="https://code.jquery.com/jquery-3.1.1.slim.min.js" integrity="sha384-A7FZj7v+d/sdmMqp/nOQwliLvUsJfDHW+k9Omg/a/EheAdgtzNs3hpfag6Ed950n" crossorigin="anonymous"></script>
                <script src="https://cdnjs.cloudflare.com/ajax/libs/tether/1.4.0/js/tether.min.js" integrity="sha384-DztdAPBWPRXSA/3eYEEUWrWCy7G5KFbe8fFjk5JAIxUYHKkDx6Qin1DkWx51bBrb" crossorigin="anonymous"></script>
                <script src="https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0-alpha.6/js/bootstrap.min.js" integrity="sha384-vBWWzlZJ8ea9aCX4pEW3rVHjgjt7zpkNpZk+02D9phzyeVkE+jo0ieGizqPLForn" crossorigin="anonymous"></script>
              </body>
            </html>
            """)
        return

    @property
    def do_POST(self):
        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={'REQUEST_METHOD': 'POST',
                     'CONTENT_TYPE': self.headers['Content-Type'],
                     })
        starting_time_dic = form["startingtime"]
        starting_time = starting_time_dic.value
        duration_dic = form["duration"]
        duration = int(duration_dic.value)
        ssid = form["ssid"].value.split(",")
        movement_head = ['WT', 'ET', 'NT', 'ST', 'WL', 'EL', 'NL', 'SL']
        # another_final_file = "finalReport_" + ssid[0] + "_6_30_360_.csv"
        excel_file_name = "bluetooth_output.xlsx"

        # start reading files for a certain date
        # ssid = ['4869']
        # dates = ['20170213', '20170214']
        dates = form["date"].value.split(",")
        # timespan = ['7:30']
        # check if the time starts as 0, since second part require starts as 0, but first part require as no 0
        if starting_time[0] == '0':
            starting_time_timespan = starting_time[1:]
        print starting_time_timespan
        timespan = [starting_time_timespan]
        # timelength = ['100']
        timelength = [duration]
        queryfield = ['Date', 'Time']
        output = ['Duration', 'Movement', 'Period']
        path = "report/result_" + str(datetime.datetime.utcnow()).replace(" ", "_").replace(":", "_").replace(".",
                                                                                                              "_") + "/"

        movement_head = ['WT', 'ET', 'NT', 'ST', 'WL', 'EL', 'NL', 'SL']

        iDate = 0
        iDuration = 1
        iWT = 2
        iET = 3
        iNT = 4
        iST = 5
        iWL = 6
        iEL = 7
        iNL = 8
        iSL = 9
        iPeriod = 10

        def RepresentsInt(s):
            try:
                int(s)
                return True
            except ValueError:
                return False

        # find files by ssid and date
        result = {}
        for strSSID in ssid:
            for strdates in dates:
                # read scv files in folder
                for infile in sorted(glob.glob('*.csv')):
                    filename = str(infile)
                    if strSSID + "-" + strdates in filename:
                        print "Current File Being Processed is: " + filename
                        reader = csv.DictReader(open(filename))

                        for row in reader:
                            for column, value in row.iteritems():
                                result.setdefault(column, []).append(value)

        # now all the potential data are save into result
        sectionRes = []
        for sid in ssid:
            itime = 0

            for time in timespan:
                dic = {}

                reportFile = path + "finalReport_" + str(sid) + '_' + str(time).replace(" ", "_").replace(":",
                                                                                                          "_").replace(
                    ".", "_") + '_' + str(timelength[itime]).replace(" ", "_").replace(":", "_").replace(".",
                                                                                                         "_") + '_.txt'

                reportFile_csv = path + "finalReport_" + str(sid) + '_' + str(time).replace(" ", "_").replace(":",
                                                                                                              "_").replace(
                    ".", "_") + '_' + str(timelength[itime]).replace(" ", "_").replace(":", "_").replace(".",
                                                                                                         "_") + '_.csv'

                if not os.path.exists(os.path.dirname(reportFile)):
                    try:
                        os.makedirs(os.path.dirname(reportFile))
                    except OSError as exc:
                        if exc.errno != errno.EEXIST:
                            raise

                if not os.path.exists(os.path.dirname(reportFile_csv)):
                    try:
                        os.makedirs(os.path.dirname(reportFile_csv))
                    except OSError as exc:
                        if exc.errno != errno.EEXIST:
                            raise

                with open(reportFile, "w") as f:
                    f.write('Date,Period,WT,ET,NT,ST,WL,EL,NL,SL\n')
                    f.close()

                with open(reportFile_csv, "wb") as fc:
                    writer = csv.writer(fc)
                    writer.writerow(['Date', 'Period', 'WT', 'ET', 'NT', 'ST', 'WL', 'EL', 'NL', 'SL'])
                    fc.close()

                for dateOpt in dates:
                    arrCounter = []
                    startdt = datetime.datetime.strptime(dateOpt + ' ' + time, '%Y%m%d %H:%M')
                    print startdt
                    enddt = startdt + datetime.timedelta(seconds=int(timelength[itime]) * 60)
                    print enddt

                    # start loop the source dictionary rows
                    datelist = result.get(queryfield[
                                              0])  # just pick up any random list to start a loop, since they all have same length
                    index = 0
                    count = 0
                    arrNumMovement = ['', '', 0, 0, 0, 0, 0, 0, 0, 0]
                    arrSumDuration = ['', '', 0, 0, 0, 0, 0, 0, 0, 0]
                    # period = ''
                    for strdate in datelist:
                        datetime_obj = datetime.datetime.strptime(strdate + ' ' + result.get(queryfield[1])[index],
                                                                  '%m/%d/%Y %H:%M:%S')

                        # print datetime_obj

                        # #################### MAIN LOOP TO DEAL WITH CELL DATA ################################
                        # check if datetime if current row fall into the range
                        if datetime_obj >= startdt and datetime_obj <= enddt:
                            # print str(datetime_obj)
                            dic.setdefault("datetime", []).append(
                                str(datetime_obj))  # add current datetime as a field
                            dic.setdefault("ssid", []).append(str(id))  # add current ssid as a field

                            # start adding user defined fields
                            arrRow = [str(dateOpt) + ' ' + str(datetime_obj), 0, 0, 0, 0, 0, 0, 0, 0, 0, '']
                            arrNumMovement[0] = str(dateOpt)
                            arrSumDuration[0] = str(dateOpt)
                            tempDuration = 0.0
                            for fieldname in output:
                                dic.setdefault(fieldname, []).append(result.get(fieldname)[index])
                                if fieldname == 'Duration':
                                    if RepresentsInt(str(result.get(fieldname)[index])):
                                        arrRow[iDuration] = result.get(fieldname)[index]
                                        tempDuration = float(result.get(fieldname)[index])
                                    else:
                                        break
                                if fieldname == 'Movement':
                                    iMove = 2
                                    for move in movement_head:
                                        # print move
                                        # print result.get(fieldname)[index]
                                        if move in result.get(fieldname)[index]:
                                            arrRow[iMove] = arrRow[iMove] + 1
                                            arrNumMovement[iMove] = arrNumMovement[iMove] + 1
                                            arrSumDuration[iMove] = arrSumDuration[iMove] + tempDuration
                                        iMove = iMove + 1
                                if fieldname == 'Period':
                                    arrRow[iPeriod] = result.get(fieldname)[index]
                                    if str(result.get(fieldname)[index]) <> '0':
                                        # period = str(result.get(fieldname)[index])
                                        arrNumMovement[1] = str(result.get(fieldname)[index])
                                        arrSumDuration[1] = str(result.get(fieldname)[index])
                            arrCounter.append(arrRow)
                        index = index + 1

                    # start writing result for a specific ssid for a specific timespan of a date
                    srcFile = path + str(sid) + '_' + str(startdt).replace(" ", "_").replace(":", "_").replace(".",
                                                                                                               "_") + '_' + str(
                        enddt).replace(" ", "_").replace(":", "_").replace(".", "_") + '_' + str(
                        datetime.datetime.utcnow()).replace(" ", "_").replace(":", "_").replace(".", "_") + ".txt"

                    if not os.path.exists(os.path.dirname(srcFile)):
                        try:
                            os.makedirs(os.path.dirname(srcFile))
                        except OSError as exc:
                            if exc.errno != errno.EEXIST:
                                raise

                    with open(srcFile, "w") as f:
                        f.write('Date,Duration,WT,ET,NT,ST,WL,EL,NL,SL,Period\n')
                        for row in arrCounter:
                            idx = 0
                            for item in row:
                                idx += 1
                                if idx == len(row):
                                    f.write(str(item) + '\n')
                                else:
                                    f.write(str(item) + ',')
                        f.close()

                    with open(reportFile, "ab") as f:
                        with open(reportFile_csv, "ab") as fc:
                            idx = 0
                            writer = csv.writer(fc)
                            arrCsvRow = []
                            for cell in arrNumMovement:
                                if idx == 0 or idx == 1:  # date and period, jsut write into the report
                                    f.write(str(cell) + ',')
                                    arrCsvRow.append(str(cell))
                                else:  # start movement section, need to devide by the num: totalDuration / numMovement
                                    # print arrSumDuration[idx]
                                    # print cell
                                    if idx == len(arrNumMovement):
                                        if cell == 0:
                                            f.write(str(cell) + '\n')
                                            arrCsvRow.append(str(cell))
                                        else:
                                            f.write(str(arrSumDuration[idx] / cell) + '\n')
                                            arrCsvRow.append(str(arrSumDuration[idx] / cell))
                                    else:
                                        if cell == 0:
                                            f.write(str(cell) + ',')
                                            arrCsvRow.append(str(cell))
                                        else:
                                            f.write(str(arrSumDuration[idx] / cell) + ',')
                                            arrCsvRow.append(str(arrSumDuration[idx] / cell))
                                idx = idx + 1
                            print arrCsvRow
                            writer.writerow(arrCsvRow)
                            f.close()
                            fc.close()
                            # count1 = 0 ==
                            # if 'WT' in dic{index}:
                            #    totalDur1 =
                            #    aveDur1 = totalDur1/count1

                # finish one dictionary for 1 specific ssid and a specific time duration
                # add it to the section result array
                # print str(dic)
                sectionRes.append(dic.copy())
                itime = itime + 1

        # start writing to report file;
        reportName = path + "report_" + str(datetime.datetime.utcnow()).replace(" ", "_").replace(":", "_").replace(
            ".", "_") + ".txt"
        if not os.path.exists(os.path.dirname(reportName)):
            try:
                os.makedirs(os.path.dirname(reportName))
            except OSError as exc:
                if exc.errno != errno.EEXIST:
                    raise

        with open(reportName, "w") as f:
            f.write(str(sectionRes))
            f.close()
        shutil.copy2(reportFile_csv, '.')  # copy file to outside folder

        # Starting for second part
        def calculate_ending_time(start_time):
            adding_hour = int(duration / 60)
            adding_minute = int(duration % 60)
            ending_time_hour = int(str(start_time).split(':')[0]) + adding_hour
            if ending_time_hour < 10:
                ending_time_hour = "0" + str(ending_time_hour)
            ending_time_minute = int(str(start_time).split(':')[1]) + adding_minute
            if ending_time_minute < 10:
                ending_time_minute = "0" + str(ending_time_minute)
            return str(ending_time_hour) + ":" + str(ending_time_minute)


        def transform_time(original_string):
            if len(original_string.split(' ')) > 2:
                old_hour_minute = original_string.split(' ', 5)[4].rsplit(':', 1)[0]
                old_hour = old_hour_minute.split(':', 1)[0]
                old_date = original_string.split(' ', 5)[3] + original_string.split(' ', 5)[1] + original_string.split(' ', 5)[
                    2]
                date = original_string[0:24]

                new_minute = int(original_string.split(' ', 5)[4].split(':', 2)[1]) + 5
                if new_minute < 10:
                    new_minute = "0" + str(new_minute)
                return str(old_hour_minute) + " - " + str(old_hour) + ":" + str(new_minute)


        def transfrom_date(original_string):
            old_hour_minute = original_string.split(' ', 5)[4].rsplit(':', 1)[0]
            old_hour = old_hour_minute.split(':', 1)[0]
            old_date = original_string.split(' ', 5)[3] + original_string.split(' ', 5)[1] + original_string.split(' ', 5)[2]
            date = original_string[0:24]
            return str(datetime.datetime.strptime(date, "%a %b %d %Y %H:%M:%S")).split(' ')[0]


        def compare_time(input_time, interval_upper_time):
            input_time_hour = str(input_time).split(':', 0)
            input_time_minute = str(input_time).split(':', 1)
            interval_upper_time_hour = str(interval_upper_time).split(':', 0)
            interval_upper_time_minute = str(interval_upper_time).split(':', 1)
            if input_time_hour > interval_upper_time_hour:
                return 0  # no need to write
                if input_time_minute > interval_upper_time_minute:
                    return 0  # no need to write
            else:
                return 1

        another_final_file = reportFile_csv
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = "DataSheet"
        date_sheet_name = sheet.title
        sheet['A1'] = 'hello world'

        result = {}
        a = ""
        b = ""
        c = 1
        d = 1
        common_array = []

        def init_final_report():
            sheet["A1"] = "Time"
            sheet["B1"] = "Period"
            sheet["C1"] = "WT"
            sheet["D1"] = "Time Interval (WT)"
            sheet["E1"] = "Travel Time (WT)"
            sheet["F1"] = "ET"
            sheet["G1"] = "Time Interval (ET)"
            sheet["H1"] = "Travel Time (ET)"
            sheet["I1"] = "NT"
            sheet["J1"] = "Time Interval (NT)"
            sheet["K1"] = "Travel Time (NT)"
            sheet["L1"] = "ST"
            sheet["M1"] = "Time Interval (ST)"
            sheet["N1"] = "Travel Time (ST)"
            sheet["O1"] = "WL"
            sheet["P1"] = "Time Interval (WL)"
            sheet["Q1"] = "Travel Time (WL)"
            sheet["R1"] = "EL"
            sheet["S1"] = "Time Interval (EL)"
            sheet["T1"] = "Travel Time (EL)"
            sheet["U1"] = "NL"
            sheet["V1"] = "Time Interval (NL)"
            sheet["W1"] = "Travel Time (NL)"
            sheet["X1"] = "SL"
            sheet["Y1"] = "Time Interval (SL)"
            sheet["Z1"] = "Travel Time (SL)"

        def fill_empty(column_name, total_row_number):
            i = 2
            while (i < total_row_number):
                if sheet[column_name + str(i)].value is None:
                    sheet[column_name + str(i)].value = sheet[column_name + str(i - 1)].value
                i += 1

        for strSSID in ssid:
            for infile in sorted(glob.glob('*.csv')):
                filename = str(infile)
                if strSSID in filename and "WT" in filename:
                    c = 1
                    print "Current File Being Processed is: " + filename
                    reader = csv.DictReader(open(filename))
                    for row in reader:
                        if len(str(row).split(',')) > 3:  # ignore missing value dirty rows
                            try:
                                for column, value in sorted(row.iteritems(), reverse=True):
                                    if column == "Time\"":
                                        if (compare_time(starting_time,
                                                         transform_time(str(value)).split(" - ")[1]) == 1 and compare_time(
                                            transform_time(str(value)).split(" - ")[0],
                                                calculate_ending_time(starting_time)) == 1):
                                            c += 1
                                            cell_name = "D" + str(c)
                                            sheet[cell_name] = transform_time(str(value))
                                            cell_previous_name = "A" + str(int(c) - 1)
                                            cell_name = "A" + str(c)
                                            if len(str(sheet[cell_previous_name].value)) > 5:  # used to compare None
                                                temp = str(sheet[cell_previous_name].value)
                                            if transfrom_date(str(value)) != temp:
                                                sheet[cell_name] = transfrom_date(str(value))
                                        else:
                                            break
                                    if column == "Strength":
                                        cell_name = "E" + str(c)
                                        sheet[cell_name] = datetime.datetime.strptime(value, '%M:%S')
                                        sheet[cell_name].number_format = "mm:ss"

                            except:
                                print("There is dirty data row in the file.")
                if strSSID in filename and "ET" in filename:
                    c = 1
                    print "Current File Being Processed is: " + filename
                    reader = csv.DictReader(open(filename))
                    for row in reader:
                        if len(str(row).split(',')) > 3:  # ignore missing value dirty rows
                            try:
                                for column, value in sorted(row.iteritems(), reverse=True):
                                    if column == "Time\"":
                                        if (compare_time(starting_time,
                                                         transform_time(str(value)).split(" - ")[1]) == 1 and compare_time(
                                            transform_time(str(value)).split(" - ")[0],
                                                calculate_ending_time(starting_time)) == 1):
                                            c += 1
                                            cell_name = "G" + str(c)
                                            sheet[cell_name] = transform_time(str(value))
                                            cell_previous_name = "A" + str(int(c) - 1)
                                            cell_name = "A" + str(c)
                                            if len(str(sheet[cell_previous_name].value)) > 5:  # used to compare None
                                                temp = str(sheet[cell_previous_name].value)
                                            if transfrom_date(str(value)) != temp:
                                                sheet[cell_name] = transfrom_date(str(value))
                                                common_array.append(c)
                                                line = c  # used to generate chart
                                        else:
                                            break
                                    if column == "Strength":
                                        cell_name = "H" + str(c)
                                        sheet[cell_name] = datetime.datetime.strptime(value, '%M:%S')
                                        sheet[cell_name].number_format = "mm:ss"
                            except:
                                print("There is dirty data row in the file.")
                if strSSID in filename and "NT" in filename:
                    c = 1
                    print "Current File Being Processed is: " + filename
                    reader = csv.DictReader(open(filename))
                    for row in reader:
                        if len(str(row).split(',')) > 3:  # ignore missing value dirty rows
                            try:
                                for column, value in sorted(row.iteritems(), reverse=True):
                                    if column == "Time\"":
                                        if (compare_time(starting_time,
                                                         transform_time(str(value)).split(" - ")[1]) == 1 and compare_time(
                                            transform_time(str(value)).split(" - ")[0],
                                                calculate_ending_time(starting_time)) == 1):
                                            c += 1
                                            cell_name = "J" + str(c)
                                            sheet[cell_name] = transform_time(str(value))
                                            cell_previous_name = "A" + str(int(c) - 1)
                                            cell_name = "A" + str(c)
                                            if len(str(sheet[cell_previous_name].value)) > 5:  # used to compare None
                                                temp = str(sheet[cell_previous_name].value)
                                            if transfrom_date(str(value)) != temp:
                                                sheet[cell_name] = transfrom_date(str(value))
                                        else:
                                            break
                                    if column == "Strength":
                                        cell_name = "K" + str(c)
                                        sheet[cell_name] = datetime.datetime.strptime(value, '%M:%S')
                                        sheet[cell_name].number_format = "mm:ss"
                            except:
                                print("There is dirty data row in the file.")
                if strSSID in filename and "ST" in filename:
                    c = 1
                    print "Current File Being Processed is: " + filename
                    reader = csv.DictReader(open(filename))
                    for row in reader:
                        if len(str(row).split(',')) > 3:  # ignore missing value dirty rows
                            try:
                                for column, value in sorted(row.iteritems(), reverse=True):
                                    if column == "Time\"":
                                        if (compare_time(starting_time,
                                                         transform_time(str(value)).split(" - ")[1]) == 1 and compare_time(
                                            transform_time(str(value)).split(" - ")[0],
                                                calculate_ending_time(starting_time)) == 1):
                                            c += 1
                                            cell_name = "M" + str(c)
                                            sheet[cell_name] = transform_time(str(value))
                                            cell_previous_name = "A" + str(int(c) - 1)
                                            cell_name = "A" + str(c)
                                            if len(str(sheet[cell_previous_name].value)) > 5:  # used to compare None
                                                temp = str(sheet[cell_previous_name].value)
                                            if transfrom_date(str(value)) != temp:
                                                sheet[cell_name] = transfrom_date(str(value))
                                        else:
                                            break
                                    if column == "Strength":
                                        cell_name = "N" + str(c)
                                        sheet[cell_name] = datetime.datetime.strptime(value, '%M:%S')
                                        sheet[cell_name].number_format = "mm:ss"
                            except:
                                print("There is dirty data row in the file.")
                if strSSID in filename and "WL" in filename:
                    c = 1
                    print "Current File Being Processed is: " + filename
                    reader = csv.DictReader(open(filename))
                    for row in reader:
                        if len(str(row).split(',')) > 3:  # ignore missing value dirty rows
                            try:
                                for column, value in sorted(row.iteritems(), reverse=True):
                                    if column == "Time\"":
                                        if (compare_time(starting_time,
                                                         transform_time(str(value)).split(" - ")[1]) == 1 and compare_time(
                                            transform_time(str(value)).split(" - ")[0],
                                                calculate_ending_time(starting_time)) == 1):
                                            c += 1
                                            cell_name = "P" + str(c)
                                            sheet[cell_name] = transform_time(str(value))
                                            cell_previous_name = "A" + str(int(c) - 1)
                                            cell_name = "A" + str(c)
                                            if len(str(sheet[cell_previous_name].value)) > 5:  # used to compare None
                                                temp = str(sheet[cell_previous_name].value)
                                            if transfrom_date(str(value)) != temp:
                                                sheet[cell_name] = transfrom_date(str(value))
                                        else:
                                            break
                                    if column == "Strength":
                                        cell_name = "Q" + str(c)
                                        sheet[cell_name] = datetime.datetime.strptime(value, '%M:%S')
                                        sheet[cell_name].number_format = "mm:ss"
                            except:
                                print("There is dirty data row in the file.")
                if strSSID in filename and "EL" in filename:
                    c = 1
                    print "Current File Being Processed is: " + filename
                    reader = csv.DictReader(open(filename))
                    for row in reader:
                        if len(str(row).split(',')) > 3:  # ignore missing value dirty rows
                            try:
                                for column, value in sorted(row.iteritems(), reverse=True):
                                    if column == "Time\"":
                                        if (compare_time(starting_time,
                                                         transform_time(str(value)).split(" - ")[1]) == 1 and compare_time(
                                            transform_time(str(value)).split(" - ")[0],
                                                calculate_ending_time(starting_time)) == 1):
                                            c += 1
                                            cell_name = "S" + str(c)
                                            sheet[cell_name] = transform_time(str(value))
                                            cell_previous_name = "A" + str(int(c) - 1)
                                            cell_name = "A" + str(c)
                                            if len(str(sheet[cell_previous_name].value)) > 5:  # used to compare None
                                                temp = str(sheet[cell_previous_name].value)
                                            if transfrom_date(str(value)) != temp:
                                                sheet[cell_name] = transfrom_date(str(value))
                                        else:
                                            break
                                    if column == "Strength":
                                        cell_name = "T" + str(c)
                                        sheet[cell_name] = datetime.datetime.strptime(value, '%M:%S')
                                        sheet[cell_name].number_format = "mm:ss"
                            except:
                                print("There is dirty data row in the file.")

                if strSSID in filename and "NL" in filename:
                    c = 1
                    print "Current File Being Processed is: " + filename
                    reader = csv.DictReader(open(filename))
                    for row in reader:
                        if len(str(row).split(',')) > 3:  # ignore missing value dirty rows
                            try:
                                for column, value in sorted(row.iteritems(), reverse=True):
                                    if column == "Time\"":
                                        if (compare_time(starting_time,
                                                         transform_time(str(value)).split(" - ")[1]) == 1 and compare_time(
                                            transform_time(str(value)).split(" - ")[0],
                                                calculate_ending_time(starting_time)) == 1):
                                            c += 1
                                            cell_name = "V" + str(c)
                                            sheet[cell_name] = transform_time(str(value))
                                            cell_previous_name = "A" + str(int(c) - 1)
                                            cell_name = "A" + str(c)
                                            if len(str(sheet[cell_previous_name].value)) > 5:  # used to compare None
                                                temp = str(sheet[cell_previous_name].value)
                                            if transfrom_date(str(value)) != temp:
                                                sheet[cell_name] = transfrom_date(str(value))
                                        else:
                                            break
                                    if column == "Strength":
                                        cell_name = "W" + str(c)
                                        sheet[cell_name] = datetime.datetime.strptime(value, '%M:%S')
                                        sheet[cell_name].number_format = "mm:ss"
                            except:
                                print("There is dirty data row in the file.")
                if strSSID in filename and "SL" in filename:
                    c = 1
                    print "Current File Being Processed is: " + filename
                    reader = csv.DictReader(open(filename))
                    for row in reader:
                        if len(str(row).split(',')) > 3:  # ignore missing value dirty rows
                            try:
                                for column, value in sorted(row.iteritems(), reverse=True):
                                    if column == "Time\"":
                                        if (compare_time(starting_time,
                                                         transform_time(str(value)).split(" - ")[1]) == 1 and compare_time(
                                            transform_time(str(value)).split(" - ")[0],
                                                calculate_ending_time(starting_time)) == 1):
                                            c += 1
                                            cell_name = "Y" + str(c)
                                            sheet[cell_name] = transform_time(str(value))
                                            cell_previous_name = "A" + str(int(c) - 1)
                                            cell_name = "A" + str(c)
                                            if len(str(sheet[cell_previous_name].value)) > 5:  # used to compare None
                                                temp = str(sheet[cell_previous_name].value)
                                            if transfrom_date(str(value)) != temp:
                                                sheet[cell_name] = transfrom_date(str(value))
                                        else:
                                            break
                                    if column == "Strength":
                                        cell_name = "Z" + str(c)
                                        sheet[cell_name] = datetime.datetime.strptime(value, '%M:%S')
                                        sheet[cell_name].number_format = "mm:ss"
                            except:
                                print("There is dirty data row in the file.")

        reader = csv.DictReader(open(another_final_file))
        i = 0
        for row in reader:
            for column, value in sorted(row.iteritems()):
                # if i < c:
                if str(column) == "Period":
                    for common_row in common_array:
                        cell_name = "B" + str(common_row)
                        sheet[cell_name] = str(value)
                if str(column) == "WT":
                    cell_name = "C" + str(common_array[i])
                    sheet[cell_name] = float(value)
                if str(column) == "ET":
                    cell_name = "F" + str(common_array[i])
                    sheet[cell_name] = float(value)
                if str(column) == "NT":
                    cell_name = "I" + str(common_array[i])
                    sheet[cell_name] = float(value)
                if str(column) == "ST":
                    cell_name = "L" + str(common_array[i])
                    sheet[cell_name] = float(value)
                if str(column) == "WL":
                    cell_name = "O" + str(common_array[i])
                    sheet[cell_name] = float(value)
                if str(column) == "EL":
                    cell_name = "R" + str(common_array[i])
                    sheet[cell_name] = float(value)
                if str(column) == "NL":
                    cell_name = "U" + str(common_array[i])
                    sheet[cell_name] = float(value)
                if str(column) == "SL":
                    cell_name = "X" + str(common_array[i])
                    sheet[cell_name] = float(value)
            i += 1

        init_final_report()
        fill_empty("B", c + 1)
        fill_empty("C", c + 1)
        fill_empty("F", c + 1)
        fill_empty("I", c + 1)
        fill_empty("L", c + 1)
        fill_empty("O", c + 1)
        fill_empty("R", c + 1)
        fill_empty("U", c + 1)
        fill_empty("X", c + 1)

        book.save(excel_file_name)

        workbook = xlrd.open_workbook(excel_file_name)
        sheets = workbook.sheets()
        wb = xlsxwriter.Workbook(excel_file_name)
        time_format = wb.add_format({'num_format': 'm:ss'})

        for sheet in sheets:  # write data from old file
            newSheet = wb.add_worksheet(sheet.name)
            for row in range(sheet.nrows):
                for col in range(sheet.ncols):
                    if (col == 4 or col == 7 or col == 10 or col == 13
                            or col == 16 or col == 19 or col == 22 or col == 25):
                        newSheet.write(row, col, sheet.cell(row, col).value, time_format)
                    else:
                        newSheet.write(row, col, sheet.cell(row, col).value)

        def create_new_chart(workbook):
            chart = workbook.add_chart({'type': 'line'})
            # Add a chart title and some axis labels.
            #chart.set_title({'name': 'Traffic Info (Date: ' + dates[rowNumber] + 'SSID: ' + ssid[0] + ')'})

            chart.set_x_axis({'name': 'Time Intervals',
                              'num_font': {
                                    # 'name': 'Arial'
                                    # 'name': 'Times New Roman',
                                    # 'regular': True,
                                    'size': 11
                                    }
                              })
            # chart.set_x_axis({'text_axis': True})
            chart.set_y_axis({'name': 'Travel Time', 'num_format': 'm:ss',
                              'num_font': {
                                  'size': 11
                              }
                              })
            chart.set_y2_axis({'name': 'Bond',
                              'num_font': {
                                  'size': 11
                              }})
            chart.set_legend({'position': 'bottom'})
            chart.set_size({'y_scale': 1.5})
            return chart

        new_chart_sheet = wb.add_worksheet("ChartSheet")
        # new_chart_sheet = xfile.get_sheet_by_name('ChartSheet')
        i = 0
        j = 2

        for common_row in common_array:  # every loop draw 8 pic
            chart = create_new_chart(wb)
            chart.set_title({'name': 'Traffic Info (SSID: ' + ssid[0] + ' Date: ' + dates[i] + ')'})
            chart.add_series({
                'name': date_sheet_name + '!$C$1',
                'categories': (date_sheet_name + '!$D$2:$D$' + str(line-1)),
                'values': (date_sheet_name + '!$C$2:$C$' + str(line-1)),
                'y2_axis': 1,
            })
            chart.add_series({
                'name': date_sheet_name + '!$E$1',
                'categories': (date_sheet_name + '!$D$2:$D$' + str(line - 1)),
                'values': (date_sheet_name + '!$E$2:$E$' + str(line-1)),
            })
            # Insert 1 the chart into the worksheet (with an offset).
            new_chart_sheet.insert_chart('B' + str(j), chart, {'x_offset': 25, 'y_offset': 10})

            chart.set_title({'name': 'Traffic Info (SSID: ' + ssid[0] + ' Date: ' + dates[i] + ')'})
            chart = create_new_chart(wb)
            chart.add_series({
                'name': date_sheet_name + '!$F$1',
                'categories': (date_sheet_name + '!$D$2:$D$' + str(line - 1)),
                'values': (date_sheet_name + '!$F$2:$F$' + str(line-1)),
                'y2_axis': 1,
            })
            chart.add_series({
                'name': date_sheet_name + '!$H$1',
                'categories': (date_sheet_name + '!$D$2:$D$' + str(line - 1)),
                'values': (date_sheet_name + '!$H$2:$H$' + str(line-1)),
            })
            # Insert 2 the chart into the worksheet (with an offset).
            new_chart_sheet.insert_chart('J' + str(j), chart, {'x_offset': 25, 'y_offset': 10})

            chart.set_title({'name': 'Traffic Info (SSID: ' + ssid[0] + ' Date: ' + dates[i] + ')'})
            chart = create_new_chart(wb)
            chart.add_series({
                'name': date_sheet_name + '!$I$1',
                'categories': (date_sheet_name + '!$D$2:$D$' + str(line - 1)),
                'values': (date_sheet_name + '!$I$2:$I$' + str(line-1)),
                'y2_axis': 1,
            })
            chart.add_series({
                'name': date_sheet_name + '!$K$1',
                'categories': (date_sheet_name + '!$D$2:$D$' + str(line - 1)),
                'values': (date_sheet_name + '!$K$2:$K$' + str(line-1)),
            })
            # Insert 3 the chart into the worksheet (with an offset).
            new_chart_sheet.insert_chart('R' + str(j), chart, {'x_offset': 25, 'y_offset': 10})

            chart.set_title({'name': 'Traffic Info (SSID: ' + ssid[0] + ' Date: ' + dates[i] + ')'})
            chart = create_new_chart(wb)
            chart.add_series({
                'name': date_sheet_name + '!$L$1',
                'categories': (date_sheet_name + '!$D$2:$D$' + str(line - 1)),
                'values': (date_sheet_name + '!$L$2:$L$' + str(line-1)),
                'y2_axis': 1,
            })
            chart.add_series({
                'name': date_sheet_name + '!$N$1',
                'categories': (date_sheet_name + '!$D$2:$D$' + str(line - 1)),
                'values': (date_sheet_name + '!$N$2:$N$' + str(line-1)),
            })
            # Insert 4 the chart into the worksheet (with an offset).
            new_chart_sheet.insert_chart('Z' + str(j), chart, {'x_offset': 25, 'y_offset': 10})

            chart.set_title({'name': 'Traffic Info (SSID: ' + ssid[0] + ' Date: ' + dates[i] + ')'})
            chart = create_new_chart(wb)
            chart.add_series({
                'name': date_sheet_name + '!$O$1',
                'categories': (date_sheet_name + '!$D$2:$D$' + str(line - 1)),
                'values': (date_sheet_name + '!$O$2:$O$' + str(line-1)),
                'y2_axis': 1,
            })
            chart.add_series({
                'name': date_sheet_name + '!$Q$1',
                'categories': (date_sheet_name + '!$D$2:$D$' + str(line - 1)),
                'values': (date_sheet_name + '!$Q$2:$Q$' + str(line-1)),
            })
            # Insert 5 the chart into the worksheet (with an offset).
            new_chart_sheet.insert_chart('AH' + str(j), chart, {'x_offset': 25, 'y_offset': 10})

            chart.set_title({'name': 'Traffic Info (SSID: ' + ssid[0] + ' Date: ' + dates[i] + ')'})
            chart = create_new_chart(wb)
            chart.add_series({
                'name': date_sheet_name + '!$R$1',
                'categories': (date_sheet_name + '!$D$2:$D$' + str(line - 1)),
                'values': (date_sheet_name + '!$R$2:$R$' + str(line-1)),
                'y2_axis': 1,
            })
            chart.add_series({
                'name': date_sheet_name + '!$T$1',
                'categories': (date_sheet_name + '!$D$2:$D$' + str(line - 1)),
                'values': (date_sheet_name + '!$T$2:$T$' + str(line-1)),
            })
            # Insert 6 the chart into the worksheet (with an offset).
            new_chart_sheet.insert_chart('AP' + str(j), chart, {'x_offset': 25, 'y_offset': 10})

            chart.set_title({'name': 'Traffic Info (SSID: ' + ssid[0] + ' Date: ' + dates[i] + ')'})
            chart = create_new_chart(wb)
            chart.add_series({
                'name': date_sheet_name + '!$U$1',
                'categories': (date_sheet_name + '!$D$2:$D$' + str(line - 1)),
                'values': (date_sheet_name + '!$U$2:$U$' + str(line-1)),
                'y2_axis': 1,
            })
            chart.add_series({
                'name': date_sheet_name + '!$W$1',
                'categories': (date_sheet_name + '!$D$2:$D$' + str(line - 1)),
                'values': (date_sheet_name + '!$W$2:$W$' + str(line-1)),
            })
            # Insert 7 the chart into the worksheet (with an offset).
            new_chart_sheet.insert_chart('AX' + str(j), chart, {'x_offset': 25, 'y_offset': 10})

            chart.set_title({'name': 'Traffic Info (SSID: ' + ssid[0] + ' Date: ' + dates[i] + ')'})
            chart = create_new_chart(wb)
            chart.add_series({
                'name': date_sheet_name + '!$X$1',
                'categories': (date_sheet_name + '!$D$2:$D$' + str(line - 1)),
                'values': (date_sheet_name + '!$X$2:$X$' + str(line-1)),
                'y2_axis': 1,
            })
            chart.add_series({
                'name': date_sheet_name + '!$Z$1',
                'categories': (date_sheet_name + '!$D$2:$D$' + str(line - 1)),
                'values': (date_sheet_name + '!$Z$2:$Z$' + str(line-1)),
            })
            # Insert 8 the chart into the worksheet (with an offset).
            new_chart_sheet.insert_chart('BF' + str(j), chart, {'x_offset': 25, 'y_offset': 10})

            i += 1
            j += 23  # for next row of pics

        wb.close()
        shutil.copy2('bluetooth_output.xlsx', path)
        # obtain current path
        dir_path = os.path.dirname(os.path.realpath(__file__))
        dir_path = dir_path + "/"
        self.send_response(200)
        self.send_header('Content-type', 'text/html')
        self.end_headers()
        self.wfile.write("<html lang=\"en\"><body>"
                            "The travel time report: "
                             "<a href = file:///" + dir_path + path + "bluetooth_output.xlsx>" + path + "bluetooth_output.xlsx " + "</a> <br/> <br/>"
                             "The bluetooth report: "
                             "<a href = file://" + dir_path + reportFile_csv + ">" + reportFile_csv + "</a>"
                         "</body></html>")
        return
server = HTTPServer(('', 8181), Handler)
server.serve_forever()
os.remove('bluetooth_output.xlsx')
os.remove(another_final_file)